"""
Escribe notas en el Excel de registro de forma segura.

Principio de diseño: NUNCA sobreescribe un valor existente en silencio. Si la
celda destino ya tiene un valor distinto al que se va a escribir, se reporta
como advertencia (para que se muestre en la pantalla de revisión) en vez de
pisarlo directamente. Esto es clave para el requisito de Carla: "que no se
equivoque ingresando datos" con carga masiva.
"""

from __future__ import annotations

from dataclasses import dataclass

import openpyxl

from .excel_parser import Estudiante, HojaDetectada, Practica


@dataclass
class ResultadoEscritura:
    ok: bool
    celda: str
    mensaje: str
    valor_anterior: object = None


def _col_letra(col: int) -> str:
    return openpyxl.utils.get_column_letter(col)


def escribir_nota(
    ws,
    hoja: HojaDetectada,
    practica: Practica,
    estudiante: Estudiante,
    campo: str,  # "C" o "I"
    valor: float,
    forzar: bool = False,
) -> ResultadoEscritura:
    """Escribe `valor` en la celda de C o I del estudiante para esa práctica.
    Si `forzar=False` (default) y ya hay un valor distinto, NO escribe --
    devuelve ok=False para que la UI lo muestre como conflicto a revisar."""

    if campo not in ("C", "I"):
        raise ValueError("campo debe ser 'C' o 'I' (P lo asigna Carla directamente, no la IA)")

    col = practica.col_c if campo == "C" else practica.col_i
    celda = ws.cell(row=estudiante.fila, column=col)
    celda_coord = f"{_col_letra(col)}{estudiante.fila}"

    anterior = celda.value
    if anterior not in (None, "") and not forzar:
        try:
            distinto = float(str(anterior).replace(",", ".")) != float(valor)
        except (ValueError, TypeError):
            distinto = True
        if distinto:
            return ResultadoEscritura(
                ok=False,
                celda=celda_coord,
                mensaje=(
                    f"{estudiante.nombre}: la celda {celda_coord} ya tiene un valor "
                    f"({anterior!r}) distinto al nuevo ({valor}). No se sobreescribe "
                    f"sin confirmación."
                ),
                valor_anterior=anterior,
            )

    celda.value = valor
    _asegurar_formula_total(ws, hoja, practica, estudiante)
    return ResultadoEscritura(ok=True, celda=celda_coord, mensaje="Escrito correctamente")


def _asegurar_formula_total(ws, hoja: HojaDetectada, practica: Practica, estudiante: Estudiante):
    """Si la columna T de esta práctica está vacía para este estudiante, la
    completa con la fórmula =C+P+I (respetando que P lo llena Carla aparte).
    Si ya tiene una fórmula (caso de las 'Tablas' de Excel con referencias
    estructuradas), no la toca -- Excel la recalcula solo."""
    if practica.col_t is None:
        return
    celda_t = ws.cell(row=estudiante.fila, column=practica.col_t)
    if celda_t.value not in (None, ""):
        return  # ya tiene fórmula o valor, no la tocamos

    col_c = _col_letra(practica.col_c)
    col_i = _col_letra(practica.col_i)
    partes = [f"{col_c}{estudiante.fila}"]
    if practica.col_p:
        partes.append(f"{_col_letra(practica.col_p)}{estudiante.fila}")
    partes.append(f"{col_i}{estudiante.fila}")
    celda_t.value = "=" + "+".join(partes)


def guardar_copia(wb, path_salida: str):
    wb.save(path_salida)
