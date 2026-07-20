"""
Parser genérico de los archivos de registro de calificaciones de Carla.

Diseño clave: NO asume celdas fijas. Busca el patrón:
  - Una fila de encabezados que contiene, en columnas consecutivas, C, P, I, T
    (o su variante numerada: C2, P2, I2, T2, C3, P3, I3, T3, ...)
  - El nombre de la práctica está en la celda inmediatamente ARRIBA de la
    columna "C" de ese bloque (puede estar fusionada sobre C:T).
  - En esa misma fila de encabezados debe existir una columna "Estudiante"
    (o similar) que marca dónde arrancan las filas de datos, y opcionalmente
    una columna "Grupo".

Esto funciona tanto si el archivo es un rango normal de celdas (ej.
"Calificaciones Labo Calor 1 2026 A.xlsx") como si es una Tabla de Excel con
fórmulas estructuradas (ej. "Calificaciones Laboratorio TQ2.xlsx",
fórmulas tipo Table[[#This Row],[C]]). En ambos casos leemos por posición
de columna/fila, no por el tipo de fórmula.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field

import openpyxl


# ---------- utilidades de texto ----------

def _norm(s) -> str:
    """Normaliza texto para comparar encabezados: sin tildes, minúsculas, sin espacios extra."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return s


STUDENT_HEADERS = {"estudiante", "nombre", "nombres", "alumno", "apellidos y nombres"}
GROUP_HEADERS = {"grupo", "grupo:"}
CODE_HEADERS = {"codigo", "código", "cod", "cedula"}

# Patrón para detectar encabezados tipo C, P, I, T, C2, P2, I2, T2...
RUBRO_RE = re.compile(r"^([cpit])(\d*)$")


@dataclass
class Practica:
    nombre: str
    header_row: int
    col_c: int
    col_p: int | None
    col_i: int
    col_t: int | None
    sufijo: str  # "" para la primera práctica, "2","3"... para las siguientes


@dataclass
class Estudiante:
    fila: int
    nombre: str
    codigo: str | None = None
    grupo: str | None = None


@dataclass
class HojaDetectada:
    hoja: str
    fila_encabezados: int
    col_estudiante: int
    col_grupo: int | None
    col_codigo: int | None
    fila_inicio_datos: int
    fila_fin_datos: int
    practicas: list[Practica] = field(default_factory=list)
    estudiantes: list[Estudiante] = field(default_factory=list)


def _find_header_rows(ws) -> list[int]:
    """Encuentra filas candidatas a ser fila de encabezados: contienen al menos
    una celda que matchea el patrón C/P/I/T y una celda tipo 'Estudiante'."""
    candidatas = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        row_vals = [(_norm(c.value)) for c in row]
        tiene_rubro = any(RUBRO_RE.match(v) for v in row_vals)
        tiene_estudiante = any(v in STUDENT_HEADERS for v in row_vals)
        if tiene_rubro and tiene_estudiante:
            candidatas.append(row[0].row)
    return candidatas


def _detect_practicas(ws, header_row: int) -> list[Practica]:
    """A partir de la fila de encabezados, agrupa columnas C/P/I/T por sufijo numérico
    y busca el nombre de la práctica en la fila inmediatamente superior."""
    # Recorremos la fila de encabezados en orden de columna. Cada nueva 'C' que
    # aparece inicia un bloque (práctica) distinto -- sin importar si el sufijo
    # está numerado (C2) o repetido tal cual (C). Esto cubre los dos estilos
    # reales que usa Carla en sus distintos archivos.
    matches = []
    for cell in ws[header_row]:
        v = _norm(cell.value)
        m = RUBRO_RE.match(v)
        if m:
            matches.append((cell.column, m.group(1), m.group(2)))
    matches.sort(key=lambda x: x[0])

    bloques: list[dict[str, int]] = []
    ultimo_sufijo = []
    for col, letra, sufijo in matches:
        if letra == "c":
            bloques.append({})
            ultimo_sufijo.append(sufijo)
        if bloques:
            bloques[-1][letra] = col

    practicas: list[Practica] = []
    for idx, letras in enumerate(bloques):
        sufijo = ultimo_sufijo[idx]
        if "c" not in letras or "i" not in letras:
            continue  # bloque incompleto, no es una práctica válida
        col_c = letras["c"]
        col_fin_bloque = max(letras.values())
        # Buscar nombre de práctica: recorrer filas hacia arriba desde header_row-1,
        # pero SOLO dentro del propio bloque de columnas (col_c..col_fin_bloque) para
        # no agarrar el título de la práctica vecina cuando están una junto a otra.
        nombre = None
        for r in range(header_row - 1, max(header_row - 6, 0), -1):
            for c in range(col_c, col_fin_bloque + 1):
                val = ws.cell(row=r, column=c).value
                if val and str(val).strip() and not RUBRO_RE.match(_norm(val)):
                    nombre = str(val).strip()
                    break
            if nombre:
                break
        practicas.append(
            Practica(
                nombre=nombre or f"Práctica {sufijo or '1'}",
                header_row=header_row,
                col_c=col_c,
                col_p=letras.get("p"),
                col_i=letras["i"],
                col_t=letras.get("t"),
                sufijo=sufijo,
            )
        )
    return practicas


def _detect_columnas_estudiante(ws, header_row: int) -> tuple[int, int | None, int | None]:
    col_estudiante = None
    col_grupo = None
    col_codigo = None
    for cell in ws[header_row]:
        v = _norm(cell.value)
        if v in STUDENT_HEADERS and col_estudiante is None:
            col_estudiante = cell.column
        elif v in GROUP_HEADERS and col_grupo is None:
            col_grupo = cell.column
        elif v in CODE_HEADERS and col_codigo is None:
            col_codigo = cell.column
    return col_estudiante, col_grupo, col_codigo


def parse_workbook(path: str) -> list[HojaDetectada]:
    """Punto de entrada: analiza todas las hojas de un archivo y devuelve
    la estructura detectada (prácticas + estudiantes) de cada una."""
    wb = openpyxl.load_workbook(path, data_only=False)
    resultados = []

    for hoja in wb.sheetnames:
        ws = wb[hoja]
        header_rows = _find_header_rows(ws)
        for header_row in header_rows:
            col_estudiante, col_grupo, col_codigo = _detect_columnas_estudiante(ws, header_row)
            if col_estudiante is None:
                continue
            practicas = _detect_practicas(ws, header_row)
            if not practicas:
                continue

            # Filas de datos: desde header_row+1 hasta que la columna estudiante quede vacía
            estudiantes = []
            r = header_row + 1
            fila_fin = header_row
            vacios_seguidos = 0
            while r <= ws.max_row and vacios_seguidos < 3:
                nombre = ws.cell(row=r, column=col_estudiante).value
                if nombre and str(nombre).strip():
                    vacios_seguidos = 0
                    grupo = ws.cell(row=r, column=col_grupo).value if col_grupo else None
                    codigo = ws.cell(row=r, column=col_codigo).value if col_codigo else None
                    estudiantes.append(
                        Estudiante(
                            fila=r,
                            nombre=str(nombre).strip(),
                            codigo=str(codigo).strip() if codigo else None,
                            grupo=str(grupo).strip() if grupo else None,
                        )
                    )
                    fila_fin = r
                else:
                    vacios_seguidos += 1
                r += 1

            if estudiantes:
                resultados.append(
                    HojaDetectada(
                        hoja=hoja,
                        fila_encabezados=header_row,
                        col_estudiante=col_estudiante,
                        col_grupo=col_grupo,
                        col_codigo=col_codigo,
                        fila_inicio_datos=header_row + 1,
                        fila_fin_datos=fila_fin,
                        practicas=practicas,
                        estudiantes=estudiantes,
                    )
                )
    return resultados


def resumen_legible(hojas: list[HojaDetectada]) -> str:
    """Genera un resumen en texto plano, pensado para mostrarse a Carla
    como el 'filtro' de qué prácticas se detectaron y cuántos estudiantes hay."""
    lineas = []
    for h in hojas:
        lineas.append(f"Hoja '{h.hoja}': {len(h.estudiantes)} estudiantes detectados")
        for p in h.practicas:
            lineas.append(f"  - Práctica: {p.nombre}  (columnas C={p.col_c}, I={p.col_i}, sufijo='{p.sufijo}')")
    return "\n".join(lineas)
